# -*- coding: utf-8 -*-
"""
Web scraping (frete/prazo) + atualização do Excel com inserção em E:F:G:H
empurrando os dados anteriores para a direita.

Requisitos (instale com pip):
    pip install selenium pandas numpy openpyxl

OBS: Você precisa ter o geckodriver/Firefox compatíveis no PATH se usar Firefox.
     Ajuste o browser se desejar.

Fluxo:
1) Carrega os links da planilha (coluna LINK de uma aba, ex.: "VIA");
2) Faz o scraping do frete/prazo para um CEP informado;
3) Insere um novo "bloco" em E:F:G:H (Frete, Prazo, Dias Úteis, URL),
   deslocando todo o conteúdo existente de E em diante 4 colunas à direita;
4) Salva o mesmo arquivo (ou, se preferir, com sufixo _ATUALIZADO).

Autor: Ezequiel Dannus
"""

import os
import re
import time
import random

import numpy as np
import pandas as pd
from datetime import datetime, timedelta
from typing import Iterable, Sequence, Optional, Tuple
from pathlib import Path
from sendemail.sendemail import enviar_email

from openpyxl import load_workbook

from selenium import webdriver
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.keys import Keys
from openpyxl.styles import Font, Alignment

data_formatada = datetime.now().strftime("%d-%m-%Y")


def calcular_dias_uteis(prazo_texto):
    meses = {
        "janeiro": 1,
        "fevereiro": 2,
        "março": 3,
        "marco": 3,
        "abril": 4,
        "maio": 5,
        "junho": 6,
        "julho": 7,
        "agosto": 8,
        "setembro": 9,
        "outubro": 10,
        "novembro": 11,
        "dezembro": 12,
        "jan": 1,
        "fev": 2,
        "mar": 3,
        "abr": 4,
        "mai": 5,
        "jun": 6,
        "jul": 7,
        "ago": 8,
        "set": 9,
        "out": 10,
        "nov": 11,
        "dez": 12,
        "/jan": 1,
        "/fev": 2,
        "/mar": 3,
        "/abr": 4,
        "/mai": 5,
        "/jun": 6,
        "/jul": 7,
        "/ago": 8,
        "/set": 9,
        "/out": 10,
        "/nov": 11,
        "/dez": 12,
    }

    dias_semana = {
        "segunda-feira": 0,
        "terça-feira": 1,
        "terca-feira": 1,
        "quarta-feira": 2,
        "quinta-feira": 3,
        "sexta-feira": 4,
        "sábado": 5,
        "sabado": 5,
        "domingo": 6,
        "segundafeira": 0,
        "terçafeira": 1,
        "tercafeira": 1,
        "quartafeira": 2,
        "quintafeira": 3,
        "sextafeira": 4,
        "segunda feira": 0,
        "terça feira": 1,
        "terca feira": 1,
        "quarta feira": 2,
        "quinta feira": 3,
        "sexta feira": 4,
    }

    hoje = datetime.now().date()
    prazo_texto = (prazo_texto or "").lower().strip()
    if not prazo_texto:
        return None

    if "amanhã" in prazo_texto:
        data_entrega = hoje + timedelta(days=1)
        dias_uteis = np.busday_count(
            np.datetime64(hoje + timedelta(days=1)),
            np.datetime64(data_entrega + timedelta(days=1)),
        )
        return int(dias_uteis)

    m = re.search(r"até\s+(\d{1,2})\s+de\s+([a-zç]+)", prazo_texto, re.IGNORECASE)
    if m:
        dia = int(m.group(1))
        mes_nome = m.group(2)
        mes = meses.get(mes_nome)
        if mes:
            ano = datetime.now().year
            try:
                data_entrega = datetime(ano, mes, dia).date()
                if data_entrega < hoje:
                    data_entrega = datetime(ano + 1, mes, dia).date()
                dias_uteis = np.busday_count(
                    np.datetime64(hoje + timedelta(days=1)),
                    np.datetime64(data_entrega + timedelta(days=1)),
                )
                return int(dias_uteis)
            except ValueError:
                pass

    dias_semana_pattern = r"\b(" + "|".join(dias_semana.keys()) + r")\b"
    dias_encontrados = re.findall(dias_semana_pattern, prazo_texto)
    if dias_encontrados:
        dia_semana_nome = dias_encontrados[0]
        dia_semana_num = dias_semana.get(dia_semana_nome)
        if dia_semana_num is not None:
            dias_ate = (dia_semana_num - hoje.weekday() + 7) % 7
            dias_ate = 7 if dias_ate == 0 else dias_ate
            data_entrega = hoje + timedelta(days=dias_ate)
            dias_uteis = np.busday_count(
                np.datetime64(hoje + timedelta(days=1)),
                np.datetime64(data_entrega + timedelta(days=1)),
            )
            return int(dias_uteis)

    return None


def type_like_human(element, text):
    element.clear()
    for char in text:
        element.send_keys(char)
        time.sleep(random.uniform(0.05, 0.2))


def setup_driver(headless):
    options = Options()
    if headless:
        options.add_argument("--headless=new")

    user_agent = "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:109.0) Gecko/20100101 Firefox/118.0"
    options.set_preference("general.useragent.override", user_agent)

    options.set_preference("dom.webdriver.enabled", False)
    options.set_preference("useAutomationExtension", False)

    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_argument("--ignore-certificate-errors")
    options.add_argument("--disable-dev-shm-usage")

    driver = webdriver.Firefox(options=options)
    driver.maximize_window()
    driver.set_page_load_timeout(60)
    return driver


def wait_any(driver, locators, timeout=10, get_text=True, visible=True):
    cond = (
        EC.visibility_of_element_located if visible else EC.presence_of_element_located
    )
    for how, sel in locators:
        try:
            el = WebDriverWait(driver, timeout).until(cond((how, sel)))
            return el.text.strip() if get_text else el
        except TimeoutException:
            continue
    return None


def scrape_data(driver, links, cep):
    linhas = []

    for i, url in enumerate(links):
        try:
            print(f"[{i+1}/{len(links)}] Acessando: {url}")
            driver.get(url)

            time.sleep(random.uniform(2.5, 5.0))

            try:
                btn_cookie = wait_any(
                    driver,
                    [
                        (
                            By.XPATH,
                            "//*[contains(@class,'cookies')]/descendant::button[1]",
                        ),
                        (
                            By.XPATH,
                            "//button[contains(., 'Aceitar') or contains(., 'Continuar') or contains(., 'Entendi')]",
                        ),
                    ],
                    timeout=3,
                    get_text=False,
                )
                if btn_cookie:
                    btn_cookie.click()
                    time.sleep(random.uniform(1, 2))
            except Exception:
                pass

            vendedor_madeira = wait_any(
                driver,
                [
                    (
                        By.XPATH,
                        "/html/body/div[1]/div[1]/main/div[2]/div[2]/div[2]/div/div[2]/div[1]/a",
                    )
                ],
                timeout=10,
            )

            if vendedor_madeira == "Madesa":
                trocar_vendedor = wait_any(
                    driver,
                    [
                        (
                            By.XPATH,
                            "/html/body/div[1]/div[1]/main/div[2]/div[2]/div[2]/div/div[1]/div[2]/div[2]",
                        )
                    ],
                    timeout=5,
                    get_text=False,
                )
                trocar_vendedor.click()
            else:
                pass

            cep_input = wait_any(
                driver,
                [
                    (By.XPATH, "//input[@id='frete']"),
                    (By.XPATH, "//input[@name='cep']"),
                    (By.XPATH, "//input[contains(@placeholder,'CEP')]"),
                    (
                        By.XPATH,
                        "//input[@type='tel' and (contains(@aria-label,'CEP') or contains(@name,'cep'))]",
                    ),
                ],
                timeout=10,
                get_text=False,
            )

            if not cep_input:

                toggle = wait_any(
                    driver,
                    [
                        (
                            By.XPATH,
                            "//*[@id='control-box-content']/div[1]/div[2]/div[2]/div/div[2]/div[9]/p/button",
                        ),
                        (
                            By.XPATH,
                            "//*[contains(.,'Alterar CEP') or contains(.,'mudar CEP')]/ancestor::button",
                        ),
                        (
                            By.XPATH,
                            "//*[@id='btnCalcularFrete' or @data-testid='calc-frete-button']",
                        ),
                    ],
                    timeout=5,
                    get_text=False,
                )
                if toggle:
                    toggle.click()
                    cep_input = wait_any(
                        driver,
                        [
                            (By.XPATH, "//input[@id='frete']"),
                            (By.XPATH, "//input[@name='cep']"),
                            (By.XPATH, "//input[contains(@placeholder,'CEP')]"),
                        ],
                        timeout=10,
                        get_text=False,
                    )

            if cep_input:

                type_like_human(cep_input, cep)
                time.sleep(random.uniform(1.0, 2.5))

                try:
                    cep_input.send_keys(Keys.RETURN)
                except Exception:
                    btn_calc = wait_any(
                        driver,
                        [
                            (By.XPATH, "//*[@id='Insira o CEP']"),
                            (
                                By.XPATH,
                                "//button[contains(.,'OK') or contains(.,'Consultar') or contains(., 'Calcular')]",
                            ),
                            (By.XPATH, "//button[@data-testid='calc-frete-button']"),
                        ],
                        timeout=3,
                        get_text=False,
                    )
                    if btn_calc:
                        btn_calc.click()
            else:
                print(f"[AVISO] Não achei campo de CEP na página: {url}")

            texto_valor_frete = wait_any(
                driver,
                [
                    (
                        By.XPATH,
                        "/html/body/div[1]/div[1]/main/div[2]/div[2]/div[2]/div/div[2]/div[9]/div[3]/div/p[2]",
                    )
                ],
                timeout=20,
            )

            texto_prazo = wait_any(
                driver,
                [
                    (By.XPATH, "//*[@class='cep-item--days css-o6edse eym5xli0']"),
                    (
                        By.XPATH,
                        "//*[contains(translate(., 'ABCDEFGHIJKLMNOPQRSTUVWXYZÁÉÍÓÚÂÊÔÃÕÇ', 'abcdefghijklmnopqrstuvwxyzáéíóúâêôãõç'), 'entrega')]",
                    ),
                ],
                timeout=20,
            )

            preco_prazo = wait_any(
                driver,
                [
                    (
                        By.XPATH,
                        "//*[@id='__next']/div/div[1]/div[2]/div/div[3]/div[4]/div/div/p/span[1]/span[2]",
                    )
                ],
                timeout=20,
            )

            preco_vista = wait_any(
                driver,
                [
                    (
                        By.XPATH,
                        "/html/body/div[1]/div[1]/main/div[2]/div[2]/div[2]/div/div[2]/div[4]/div/div[2]/div/div[1]/span",
                    )
                ],
                timeout=20,
            )

            dias_uteis = calcular_dias_uteis(texto_prazo)

            linhas.append(
                {
                    "Frete": texto_valor_frete,
                    "Prazo": dias_uteis,
                    "P_prazo": preco_prazo,
                    "P_vista": preco_vista,
                }
            )
            print(
                f"[OK] URL: {url}\n     CEP: {cep} | Frete: {texto_valor_frete} | Prazo: {dias_uteis}, Preço/Prazo: {preco_prazo}, A vista: {preco_vista}\n"
            )

        except Exception as e:
            print(f"[ERRO] {url} -> {e}")
            linhas.append(
                {
                    "URL": url,
                    "CEP": cep,
                    "Frete": "Erro",
                    "Prazo": "-",
                    "Dias Úteis": "-",
                }
            )
        finally:

            driver.delete_all_cookies()

    return pd.DataFrame(linhas)


def inserir_bloco_efgh(
    xlsx_path: str,
    sheet: Optional[str],
    novos_dados: Iterable[Tuple[object, object, object, object]],
    tem_cabecalho: bool = True,
    salvar_em: Optional[str] = None,
) -> str:
    wb = load_workbook(xlsx_path)
    ws = wb[sheet] if sheet else wb.active

    ws.insert_cols(5, amount=4)

    if tem_cabecalho:

        data_hoje = datetime.now().strftime("%d/%m/%Y")

        headers = [
            f"Frete ({data_hoje})",
            f"Prazo ({data_hoje})",
            f"Dias Uteis ({data_hoje})",
            f"URL ({data_hoje})",
        ]

        for col_idx, header_text in enumerate(headers, start=5):
            cell = ws.cell(row=1, column=col_idx, value=header_text)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    start_row = 2 if tem_cabecalho else 1
    for i, linha in enumerate(novos_dados, start=start_row):
        if not isinstance(linha, Sequence) or len(linha) != 4:
            raise ValueError(
                f"Cada linha deve ter 4 valores (E,F,G,H). Recebido: {linha!r}"
            )
        e, f, g, h = linha
        ws.cell(row=i, column=5, value=e)  # E
        ws.cell(row=i, column=6, value=f)  # F
        ws.cell(row=i, column=7, value=g)  # G
        ws.cell(row=i, column=8, value=h)  # H

    if salvar_em is None:
        p = Path(xlsx_path)
        out = p
    else:
        out = Path(salvar_em)
    wb.save(out)
    return str(out)


def load_links(file_path, sheet="MADEIRA"):
    s = pd.read_excel(file_path, sheet_name=sheet)["LINK"]
    links = s.dropna().astype(str).str.strip().replace({"": np.nan}).dropna().tolist()
    return links


def main():

    input_xlsx = r"C:\Users\ezequiel\webScraping- WEB\Conferências OST.xlsx"
    sheet_links = "MADEIRA"
    cep = "01449010"
    headless = False

    links = load_links(input_xlsx, sheet=sheet_links)
    if not links:
        print("Nenhum link válido encontrado.")
        return

    driver = setup_driver(headless=headless)
    try:
        df = scrape_data(driver, links, cep)
    finally:
        driver.quit()

    out_dir = os.path.dirname(input_xlsx) or "."
    relatorio = os.path.join(out_dir, f"resultados_{data_formatada}.xlsx-Madeira")
    df.to_excel(relatorio, index=False)

    print(f"\nBloco inserido em E:F:G:H na aba '{sheet_links}'.")
    print(f"Relatório do scraping salvo em: {relatorio}")


if __name__ == "__main__":
    main()
    enviar_email(
        f"resultados_{data_formatada}-Madeira.xlsx",
        "ezequiel@madesa.com",
        "v.ost@madesa.com",
        "Relatório de Frete e Prazo - Madeira",
    )
