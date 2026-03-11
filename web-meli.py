# -*- coding: utf-8 -*-
"""
Web scraping (frete/prazo) + atualização do Excel com inserção em E:F:G:H
empurrando os dados anteriores para a direita.

Requisitos (instale com pip):
    pip install selenium pandas numpy openpyxl

OBS: Você precisa ter o geckodriver/Firefox compatíveis no PATH se usar Firefox.
     Ajuste o browser se desejar.

Fluxo:
1) Carrega os links da planilha (coluna LINK de uma aba, ex.: "VIA");
2) Faz o scraping do frete/prazo para um CEP informado;
3) Insere um novo "bloco" em E:F:G:H (Frete, Prazo, Dias Úteis, URL),
   deslocando todo o conteúdo existente de E em diante 4 colunas à direita;
4) Salva o mesmo arquivo (ou, se preferir, com sufixo _ATUALIZADO).

Autor: Ezequiel Dannus
"""

import os
import re
import time

import numpy as np
import pandas as pd
from datetime import datetime, timedelta
from typing import Iterable, Sequence, Optional, Tuple
from pathlib import Path

from openpyxl import load_workbook

from selenium import webdriver
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    TimeoutException,
    WebDriverException,
    NoSuchElementException,
)
from selenium.webdriver.common.keys import Keys


# -------------------- PARSERS --------------------
def calcular_dias_uteis(prazo_texto):
    meses = {
        "janeiro": 1,
        "fevereiro": 2,
        "março": 3,
        "marco": 3,
        "abril": 4,
        "maio": 5,
        "junho": 6,
        "julho": 7,
        "agosto": 8,
        "setembro": 9,
        "outubro": 10,
        "novembro": 11,
        "dezembro": 12,
        "jan": 1,
        "fev": 2,
        "mar": 3,
        "abr": 4,
        "mai": 5,
        "jun": 6,
        "jul": 7,
        "ago": 8,
        "set": 9,
        "out": 10,
        "nov": 11,
        "dez": 12,
        "/jan": 1,
        "/fev": 2,
        "/mar": 3,
        "/abr": 4,
        "/mai": 5,
        "/jun": 6,
        "/jul": 7,
        "/ago": 8,
        "/set": 9,
        "/out": 10,
        "/nov": 11,
        "/dez": 12,
    }

    dias_semana = {
        "segunda-feira": 0,
        "terça-feira": 1,
        "terca-feira": 1,
        "quarta-feira": 2,
        "quinta-feira": 3,
        "sexta-feira": 4,
        "sábado": 5,
        "sabado": 5,
        "domingo": 6,
        "segundafeira": 0,
        "terçafeira": 1,
        "tercafeira": 1,
        "quartafeira": 2,
        "quintafeira": 3,
        "sextafeira": 4,
        "segunda feira": 0,
        "terça feira": 1,
        "terca feira": 1,
        "quarta feira": 2,
        "quinta feira": 3,
        "sexta feira": 4,
    }

    hoje = datetime.now().date()
    prazo_texto = (prazo_texto or "").lower().strip()
    if not prazo_texto:
        return None

    if "amanhã" in prazo_texto:
        data_entrega = hoje + timedelta(days=1)
        dias_uteis = np.busday_count(
            np.datetime64(hoje + timedelta(days=1)),
            np.datetime64(data_entrega + timedelta(days=1)),
        )
        return int(dias_uteis)

    # até 12 de agosto
    m = re.search(r"até\s+(\d{1,2})\s+de\s+([a-zç]+)", prazo_texto, re.IGNORECASE)
    if m:
        dia = int(m.group(1))
        mes_nome = m.group(2)
        mes = meses.get(mes_nome)
        if mes:
            ano = datetime.now().year
            try:
                data_entrega = datetime(ano, mes, dia).date()
                if data_entrega < hoje:
                    data_entrega = datetime(ano + 1, mes, dia).date()
                dias_uteis = np.busday_count(
                    np.datetime64(hoje + timedelta(days=1)),
                    np.datetime64(data_entrega + timedelta(days=1)),
                )
                return int(dias_uteis)
            except ValueError:
                pass

    # “terça-feira”, etc.
    dias_semana_pattern = r"\b(" + "|".join(dias_semana.keys()) + r")\b"
    dias_encontrados = re.findall(dias_semana_pattern, prazo_texto)
    if dias_encontrados:
        dia_semana_nome = dias_encontrados[0]
        dia_semana_num = dias_semana.get(dia_semana_nome)
        if dia_semana_num is not None:
            dias_ate = (dia_semana_num - hoje.weekday() + 7) % 7
            dias_ate = 7 if dias_ate == 0 else dias_ate
            data_entrega = hoje + timedelta(days=dias_ate)
            dias_uteis = np.busday_count(
                np.datetime64(hoje + timedelta(days=1)),
                np.datetime64(data_entrega + timedelta(days=1)),
            )
            return int(dias_uteis)

    return None


def parse_frete_e_prazo(texto_valor, texto_prazo):
    # valor
    valor_frete = "Sem buybox"
    if texto_valor:
        if "grátis" in texto_valor.lower():
            valor_frete = "GRÁTIS"
        else:
            m = re.search(r"R\$[\s]?([\d\.,]+)", texto_valor)
            if m:
                valor_frete = m.group(1).strip()

    # prazo (tenta normalizar um pouco)
    prazo_final = "-"
    if texto_prazo:
        # tenta capturar “10 - 12 de agosto” ou “12 de agosto”
        m_int = re.search(
            r"(\d{1,2})\s*-\s*(\d{1,2})\s*de\s*([A-Za-zÀ-ú]+)",
            texto_prazo,
            re.IGNORECASE,
        )
        if m_int:
            prazo_final = f"{m_int.group(2)} de {m_int.group(3)}"
        else:
            m_sim = re.search(
                r"(\d{1,2})\s*de\s*([A-Za-zÀ-ú]+)", texto_prazo, re.IGNORECASE
            )
            if m_sim:
                prazo_final = f"{m_sim.group(1)} de {m_sim.group(2)}"
            else:
                prazo_final = texto_prazo.strip()

    dias_uteis = calcular_dias_uteis(prazo_final)
    dias_uteis = dias_uteis if dias_uteis is not None else "-"

    return valor_frete, prazo_final, dias_uteis


# -------------------- SELENIUM --------------------
def setup_driver(headless):
    options = Options()
    if headless:
        options.add_argument("--headless=new")
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_argument("--ignore-certificate-errors")

    options.add_argument("--disable-dev-shm-usage")
    options.add_argument(
        "--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/139.0.0.0 Safari/537.36"
    )
    driver = webdriver.Firefox(options=options)
    driver.maximize_window()
    driver.set_page_load_timeout(60)
    return driver


def wait_any(driver, locators, timeout=10, get_text=True, visible=True):
    """Tenta vários XPaths/CSS; retorna o primeiro texto/elemento encontrado."""
    cond = (
        EC.visibility_of_element_located if visible else EC.presence_of_element_located
    )
    for how, sel in locators:
        try:
            el = WebDriverWait(driver, timeout).until(cond((how, sel)))
            return el.text.strip() if get_text else el
        except TimeoutException:
            continue
    return None


def clear_and_type(el, text):
    el.clear()
    el.send_keys(text)


def scrape_data(driver, links, cep):
    linhas = []

    for url in links:
        try:
            driver.get(url)

            # (opcional) fecha banner de cookies se aparecer
            try:
                btn_cookie = wait_any(
                    driver,
                    [
                        (
                            By.XPATH,
                            "//*[contains(@class,'cookies')]/descendant::button[1]",
                        ),
                        (
                            By.XPATH,
                            "//button[contains(., 'Aceitar') or contains(., 'Continuar')]",
                        ),
                    ],
                    timeout=3,
                    get_text=False,
                )
                if btn_cookie:
                    btn_cookie.click()
            except Exception:
                pass

            # Localiza campo de CEP (testa algumas possibilidades)
            cep_input = wait_any(
                driver,
                [
                    (By.XPATH, "//input[@id='frete']"),
                    (By.XPATH, "//input[@name='cep']"),
                    (By.XPATH, "//input[contains(@placeholder,'CEP')]"),
                    (
                        By.XPATH,
                        "//input[@type='tel' and (contains(@aria-label,'CEP') or contains(@name,'cep'))]",
                    ),
                ],
                timeout=10,
                get_text=False,
            )

            if not cep_input:
                # algumas páginas exigem clicar em “Calcular frete”/“Alterar CEP” antes
                toggle = wait_any(
                    driver,
                    [
                        (
                            By.XPATH,
                            "//*[contains(.,'Consultar') and contains(.,'frete')]/ancestor::button",
                        ),
                        (
                            By.XPATH,
                            "//*[contains(.,'Alterar CEP') or contains(.,'mudar CEP')]/ancestor::button",
                        ),
                        (
                            By.XPATH,
                            "//*[@id='btnCalcularFrete' or @data-testid='calc-frete-button']",
                        ),
                    ],
                    timeout=5,
                    get_text=False,
                )
                if toggle:
                    toggle.click()
                    cep_input = wait_any(
                        driver,
                        [
                            (By.XPATH, "//input[@id='frete']"),
                            (By.XPATH, "//input[@name='cep']"),
                            (By.XPATH, "//input[contains(@placeholder,'CEP')]"),
                        ],
                        timeout=10,
                        get_text=False,
                    )

            if cep_input:
                clear_and_type(cep_input, cep)
                # tenta apertar Enter ou clicar no botão “OK/Calcular”
                try:
                    time.sleep(3)
                    cep_input.send_keys(Keys.RETURN)
                except Exception:
                    btn_calc = wait_any(
                        driver,
                        [
                            (By.XPATH, "//*[@id='cep-box']/div[1]/div/form/div/button"),
                            (
                                By.XPATH,
                                "//button[contains(.,'OK') or contains(.,'Consultar')]",
                            ),
                        ],
                        timeout=3,
                        get_text=False,
                    )
                    if btn_calc:
                        btn_calc.click()
            else:
                print(f"[AVISO] Não achei campo de CEP na página: {url}")

            # Aguarda bloco de frete aparecer; depois lê valor e prazo
            texto_valor = wait_any(
                driver,
                [
                    (By.XPATH, "//*[@class='cep-item--value css-d17uor eym5xli0']"),
                    (
                        By.XPATH,
                        "//*[contains(@class,'frete') and (contains(.,'R$') or contains(.,'Grátis'))]",
                    ),
                ],
                timeout=15,
            )

            texto_prazo = wait_any(
                driver,
                [
                    (By.XPATH, "//*[@class='cep-item--days css-o6edse eym5xli0']"),
                    (
                        By.XPATH,
                        "//*[contains(translate(., 'ABCDEFGHIJKLMNOPQRSTUVWXYZÁÉÍÓÚÂÊÔÃÕÇ', "
                        "'abcdefghijklmnopqrstuvwxyzáéíóúâêôãõç'), 'entrega')]",
                    ),
                ],
                timeout=15,
            )

            valor_frete, prazo_final, dias_uteis = parse_frete_e_prazo(
                texto_valor, texto_prazo
            )

            linhas.append(
                {
                    "URL": url,
                    "CEP": cep,
                    "Frete": valor_frete,
                    "Prazo": prazo_final,
                    "Dias Úteis": dias_uteis,
                }
            )
            driver.delete_all_cookies()

            print(
                f"[OK] URL: {url}\n     CEP: {cep} | Frete: {valor_frete} | Prazo: {prazo_final} ({dias_uteis})"
            )

        except Exception as e:
            print(f"[ERRO] {url} -> {e}")
            linhas.append(
                {
                    "URL": url,
                    "CEP": cep,
                    "Frete": "Erro",
                    "Prazo": "-",
                    "Dias Úteis": "-",
                }
            )
            driver.delete_all_cookies()

    return pd.DataFrame(linhas)


# -------------------- EXCEL (inserção E:F:G:H) --------------------
def inserir_bloco_efgh(
    xlsx_path: str,
    sheet: Optional[str],
    novos_dados: Iterable[Tuple[object, object, object, object]],
    tem_cabecalho: bool = True,
    salvar_em: Optional[str] = None,
) -> str:
    """
    Insere um novo bloco nas colunas E:F:G:H, empurrando tudo que está da E para a direita.

    novos_dados: iterável de tuplas (E, F, G, H) na ordem das linhas..y
    """
    wb = load_workbook(xlsx_path)
    ws = wb[sheet] if sheet else wb.active

    # 1) Inserir 4 colunas a partir da coluna E (índice 5)
    ws.insert_cols(5, amount=4)

    # 2) Escrever os novos dados nas colunas E..H
    start_row = 2 if tem_cabecalho else 1
    for i, linha in enumerate(novos_dados, start=start_row):
        if not isinstance(linha, Sequence) or len(linha) != 4:
            raise ValueError(
                f"Cada linha deve ter 4 valores (E,F,G,H). Recebido: {linha!r}"
            )
        e, f, g, h = linha
        ws.cell(row=i, column=5, value=e)  # E
        ws.cell(row=i, column=6, value=f)  # F
        ws.cell(row=i, column=7, value=g)  # G
        ws.cell(row=i, column=8, value=h)  # H

    # 3) Salvar
    if salvar_em is None:
        p = Path(xlsx_path)
        out = p  # salva por cima; troque por p.with_name(p.stem + "_ATUALIZADO.xlsx") se preferir
    else:
        out = Path(salvar_em)
    wb.save(out)
    return str(out)


# -------------------- UTIL --------------------
def load_links(file_path, sheet="MELI"):
    s = pd.read_excel(file_path, sheet_name=sheet)["LINK"]
    links = (
        s.dropna()
        .astype(str)
        .str.strip()
        .replace({"": np.nan})
        .dropna()
        .tolist()  # mantém a ordem da planilha
    )
    return links


# -------------------- MAIN --------------------
def main():
    # === Ajuste caminhos/parametros ao seu ambiente ===
    input_xlsx = r"C:\Users\ezequiel\webScraping- WEB\Conferências OST.xlsx"
    sheet_links = (
        "MADEIRA"  # aba onde existe a coluna LINK e onde vamos inserir E:F:G:H
    )
    tem_cabecalho = True  # True se a linha 1 é cabeçalho
    cep = "01449010"
    headless = False  # coloque True para rodar sem abrir janela

    # 1) Carregar links mantendo a mesma ordem da planilha
    links = load_links(input_xlsx, sheet=sheet_links)
    if not links:
        print("Nenhum link válido encontrado.")
        return

    # 2) Scraping
    driver = setup_driver(headless=headless)
    try:
        df = scrape_data(driver, links, cep)
    finally:
        driver.quit()

    # 3) Preparar dados para E:F:G:H-
    #    Mapeamento: E=Frete, F=Prazo, G=Dias Úteis, H=URL
    #    Garante mesmo número de linhas que a quantidade de links lidos.
    #    Se o scraping falhar em algumas, ainda assim haverá uma linha "Erro".

    # (opcional) também salvar um relatório à parte por data
    data_formatada = datetime.now().strftime("%d-%m-%Y")
    out_dir = os.path.dirname(input_xlsx) or "."
    relatorio = os.path.join(out_dir, f"resultados_{data_formatada}.xlsx")
    df.to_excel(relatorio, index=False)

    print(f"\nBloco inserido em E:F:G:H na aba '{sheet_links}'.")
    print(f"Relatório do scraping salvo em: {relatorio}")


if __name__ == "__main__":
    main()
